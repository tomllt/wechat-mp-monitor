#!/usr/bin/env python3
"""
将微信公众号清单 Excel 转换为 JSON 格式
用法: python3 convert-excel-to-json.py
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("⚠️  缺少 openpyxl 库，正在安装...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

def convert_excel_to_json(excel_path: str, output_path: str):
    """将 Excel 转换为 JSON"""
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    accounts = []
    
    # 跳过前两行（第1行是标题，第2行是列名）
    # 列名: 排名, 企业名称, 企业类型, 公众号名称, 服务号名称, 小程序名称
    for row_idx in range(3, sheet.max_row + 1):
        # 按列号访问: A=1, B=2, C=3, D=4, E=5, F=6
        rank_val = sheet.cell(row=row_idx, column=1).value
        rank = int(rank_val) if rank_val and rank_val != 'None' else None
        
        account = {
            "rank": rank,
            "unitName": str(sheet.cell(row=row_idx, column=2).value) if sheet.cell(row=row_idx, column=2).value else "",
            "type": str(sheet.cell(row=row_idx, column=3).value) if sheet.cell(row=row_idx, column=3).value else "",
            "officialAccount": str(sheet.cell(row=row_idx, column=4).value) if sheet.cell(row=row_idx, column=4).value and sheet.cell(row=row_idx, column=4).value != 'None' else None,
            "serviceAccount": str(sheet.cell(row=row_idx, column=5).value) if sheet.cell(row=row_idx, column=5).value and sheet.cell(row=row_idx, column=5).value != 'None' else None,
            "miniProgram": str(sheet.cell(row=row_idx, column=6).value) if sheet.cell(row=row_idx, column=6).value and sheet.cell(row=row_idx, column=6).value != 'None' else None,
        }
        
        accounts.append(account)
    
    # 写入 JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 转换完成: {len(accounts)} 个公众号")
    print(f"   输出文件: {output_path}")
    
    # 统计信息
    has_official = sum(1 for a in accounts if a['officialAccount'])
    print(f"   有公众号名称: {has_official} 个")
    print(f"   企业类型分布:")
    types = {}
    for a in accounts:
        t = a['type'] or '未知'
        types[t] = types.get(t, 0) + 1
    for t, count in types.items():
        print(f"     - {t}: {count}")

if __name__ == '__main__':
    script_dir = Path(__file__).parent
    excel_path = script_dir / '微信公众号清单.xlsx'
    output_path = script_dir / '微信公众号清单.json'
    
    convert_excel_to_json(str(excel_path), str(output_path))
